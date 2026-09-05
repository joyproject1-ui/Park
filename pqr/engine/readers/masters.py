# -*- coding: utf-8 -*-
"""적격성 마스터파일 — 10.2 제조설비, 10.3~10.5 제조지원 설비의 IQ/OQ/PQ 문서번호·완료일."""
import re
from openpyxl import load_workbook

DOC = re.compile(r"^(DQ|IQ|OQ|PQ|IOQ|OPQ|QM)\d*")
DATE = re.compile(r"\d{4}[.\-/]\s?\d{1,2}[.\-/]\s?\d{1,2}")


def _cell(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def _kind(doc):
    m = re.match(r"^([A-Z]+)", doc or "")
    return m.group(1) if m else ""


def _plain(text):
    return re.sub(r"\s+", "", text or "")


def _sheets(path, preferred):
    """정해 둔 시트를 먼저, 그다음 나머지 시트 — 회사가 서식을 개정하면 시트 이름이 바뀐다.

    2026-09 디겐타 자료의 마스터파일은 시트가 '생산장비' 였다(예전 '제조시설적격성평가현황').
    이름을 정해 두면 개정될 때마다 그 항이 통째로 비므로, 머리행을 보고 고른다.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    order = ([preferred] if preferred in wb.sheetnames else []) + \
            [n for n in wb.sheetnames if n != preferred]
    for name in order:
        yield name, list(wb[name].iter_rows(values_only=True))


def _header(rows, test, limit=12):
    """머리행 번호와 그 칸 글자. 없으면 (None, None)."""
    for i, row in enumerate(rows[:limit]):
        cells = [_cell(v) for v in row]
        if test(cells):
            return i, cells
    return None, None


def _is_equipment_header(cells):
    joined = _plain(" ".join(cells))
    return "관리번호" in cells and "보고서문서번호" in joined


def _is_support_header(cells):
    return "관리번호" in cells and any("IQ 문서번호" in c for c in cells)


def equipment_docs(path, sheet="제조시설적격성평가현황"):
    """{관리번호: {"name": 장비명, "line": 라인, "docs": [(문서번호, 완료일), ...]}}

    마스터파일은 장비 한 대가 여러 행(연번 행 + 이어지는 행)이라, 관리번호가 나온 뒤
    관리번호 칸이 빈 행은 같은 장비의 문서로 본다.
    """
    rows = header = col = None
    for _name, sheet_rows in _sheets(path, sheet):
        header, cells = _header(sheet_rows, _is_equipment_header)
        if header is not None:
            rows = sheet_rows
            col = {name: cells.index(name) for name in cells if name}
            break
    if header is None:
        raise ValueError("설비 마스터파일의 머리행(관리번호·보고서 문서번호)을 찾지 못했습니다.")
    c_id = col["관리번호"]
    c_doc = col[next(k for k in col if "문서 번호" in k or "문서번호" in k)]
    # 완료일은 개정되며 '승인일자' 로 이름이 바뀌었다 — 둘 다 받는다.
    c_date = next((col[k] for k in ("완료일", "승인일자", "승인일") if k in col), None)
    c_name = next((col[k] for k in ("장비명", "설비명") if k in col), None)
    c_line = col.get("라인")
    out, current = {}, None
    for row in rows[header + 1:]:
        cells = list(row) + [None] * 25
        mid = _cell(cells[c_id])
        if mid:
            current = mid
            out.setdefault(current, {"name": _cell(cells[c_name]) if c_name is not None else "",
                                     "line": _cell(cells[c_line]) if c_line is not None else "", "docs": []})
        if current is None:
            continue
        doc = _cell(cells[c_doc])
        if doc and DOC.match(doc):
            date = _cell(cells[c_date]) if c_date is not None else ""
            out[current]["docs"].append((doc, date))
    return out


def latest_by_kind(docs):
    """[(doc, date)] → {"IQ": (doc, date), "OQ": …, "PQ": …} 가장 최근 완료일 기준 (같은 종류 여러 건이면 전부)."""
    by = {}
    for doc, date in docs:
        by.setdefault(_kind(doc), []).append((doc, date))
    return by


# 마스터파일의 '사유' 칸은 그 적격성평가가 어디에 적용되는지를 적어 둔다.
OTHER_LINE = re.compile(r"액제|점안|내용액|주사|캡슐|정제|패취|좌제")
PLACE = re.compile(r"\d+\s*[층실]|[가-힣]+[실층]")


def applies(why, line="연고"):
    """이 사유의 적격성평가를 이 제품 보고서에 실어도 되는가.

    '액제 라인 리모델링에만 적용 (23)' 이나 '3층 예비실, 입고대기실 리모델링 (21)' 처럼
    다른 라인·다른 방의 공사는 싣지 않는다(담당자 2026-09: "액제라인만 23년도에 공사한
    경우는 표시했으니 잘 참고해서 작성해"). 사유가 비었거나 우리 라인을 짚었으면 싣는다.
    """
    text = (why or "").strip()
    if not text:
        return True
    if line and line in text:
        return True
    return not (OTHER_LINE.search(text) or PLACE.search(text))


def support_docs(path, sheet="제조지원 설비 & IT 시스템"):
    """{관리번호(앞 토큰): {"name": 설비명, "system": 시스템, "IQ": [(doc, date)], "OQ": …, "PQ": …, "DQ": …}}"""
    rows = header = col = head = None
    for _name, sheet_rows in _sheets(path, sheet):
        header, cells = _header(sheet_rows, _is_support_header)
        if header is not None:
            rows = sheet_rows
            head = cells
            col = {name: j for j, name in enumerate(cells) if name}
            break
    if header is None:
        # 개정된 마스터파일은 제조지원설비도 제조설비와 같은 꼴이다(관리번호·보고서 문서번호·승인일자).
        # 읽지 못했다고 물러나면 10.3~10.5 가 통째로 빈다 — 같은 값을 그 꼴에서 읽어 온다.
        return _support_from_equipment(path)
    # 종류마다 '사유' 칸이 하나씩 있고 이름이 모두 같다 — 이름 사전으로는 하나로 겹치므로
    # 머리행에서 바로 짚는다. 승인일 칸 다음에 오는 '사유' 가 그 종류의 것이다.
    why_col = {}
    for kind in ("DQ", "IQ", "OQ", "PQ"):
        acol = next((col[k] for k in col if k.startswith(kind + " 승인")), None)
        if acol is None:
            continue
        after = next((j for j in range(acol + 1, len(head)) if head[j]), None)
        if after is not None and head[after].startswith("사유"):
            why_col[kind] = after
    out = {}
    system = ""
    for row in rows[header + 1:]:
        cells = list(row) + [None] * 25
        sysname = _cell(cells[col["시스템"]]) if "시스템" in col else ""
        if sysname:
            system = sysname
        raw_id = _cell(cells[col["관리번호"]])
        if not raw_id:
            continue
        key = raw_id.split()[0]
        entry = {"name": _cell(cells[col.get("설비명", 0)]), "system": system, "raw_id": raw_id,
                 "why": {}}
        for kind in ("DQ", "IQ", "OQ", "PQ"):
            dcol = next((col[k] for k in col if k.startswith(kind + " 문서")), None)
            acol = next((col[k] for k in col if k.startswith(kind + " 승인")), None)
            docs = [d for d in re.split(r"\s+", _cell(cells[dcol])) if d] if dcol is not None else []
            dates = DATE.findall(_cell(cells[acol])) if acol is not None else []
            # 사유는 줄바꿈으로 나뉘어 문서번호와 차례가 맞는다 ('최초 제정 (20)' / '액제 라인 …에만 적용 (23)').
            wcol = why_col.get(kind)
            whys = [w.strip() for w in str(row[wcol] or "").split("\n")] if (
                wcol is not None and wcol < len(row)) else []
            entry[kind] = [(docs[i] if i < len(docs) else "", dates[i] if i < len(dates) else "")
                           for i in range(max(len(docs), len(dates)))]
            entry["why"][kind] = [whys[i] if i < len(whys) else "" for i in range(len(entry[kind]))]
        out[key] = entry
    return out


def _support_from_equipment(path):
    """제조설비 꼴(관리번호·보고서 문서번호·승인일자)로 적힌 제조지원설비 마스터파일을 읽는다."""
    out = {}
    for key, entry in equipment_docs(path, sheet=None).items():
        by = latest_by_kind(entry["docs"])
        # 원본 문서 목록도 남긴다 — PQ 마스터의 보고서는 QM… 으로 적혀 종류로 갈리지 않는다.
        row = {"name": entry.get("name", ""), "system": entry.get("line", ""), "raw_id": key,
               "why": {}, "docs": list(entry["docs"])}
        for kind in ("DQ", "IQ", "OQ", "PQ"):
            row[kind] = by.get(kind, [])
            row["why"][kind] = [""] * len(row[kind])
        out[key.split()[0] if key.split() else key] = row
    return out


def pv_by_code(path, code, sheet=None):
    """PV 마스터에서 문서 코드(QUIO3 · QUIO2 …)가 든 계획/보고서 묶음을 제품명과 상관없이 찾는다.

    [{plan, reason, kind, report, report_date, lots:[(seq, lot, mfg)], revalidation}] — 계획 행 뒤에
    Lot 행(1st·2nd·3rd)이 이어진다.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    rows = [[_cell(v) for v in r] for r in ws.iter_rows(values_only=True)]
    header = None
    for i, r in enumerate(rows[:15]):
        if any("Lot" in c or "제조번호" in c for c in r) and any("계획" in c or "PV" in c or "Plan" in c for c in r):
            header = i; break
    out, current = [], None
    for r in rows[(header + 1) if header is not None else 0:]:
        r = r + [""] * 14
        plan_col = next((i for i, c in enumerate(r) if re.match(r"^PV\d{2}-.*-P", c)), None)
        if plan_col is not None:
            plan = r[plan_col]
            if code not in plan.upper().replace("0", "O"):
                current = None; continue
            rep_col = next((i for i in range(plan_col + 1, len(r)) if re.match(r"^PV\d{2}-.*-R", r[i])), None)
            rep = r[rep_col] if rep_col is not None else ""
            m = re.search(r"(\d{4}\.\d{2}\.\d{2})", rep)
            current = {"plan": plan.split("(")[0].strip(), "reason": r[plan_col + 1], "kind": r[plan_col + 2],
                       "report": re.sub(r"\s*\(\d{4}\.\d{2}\.\d{2}\)?\s*$", "", rep).strip(),
                       "report_date": m.group(1) if m else "", "lots": [],
                       "revalidation": r[rep_col + 1] if rep_col is not None else ""}
            out.append(current)
        if current is None:
            continue
        seq_col = next((i for i, c in enumerate(r) if re.match(r"^\d(st|nd|rd|th)$", c)), None)
        if seq_col is not None and r[seq_col + 1]:
            current["lots"].append((r[seq_col], r[seq_col + 1], r[seq_col + 2]))
    return out
