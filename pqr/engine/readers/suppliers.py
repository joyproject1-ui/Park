# -*- coding: utf-8 -*-
"""공급업체 목록(8.1.1 원료 · 8.1.3 자재) 과 주성분 공급망 마스터파일(8.1.2)."""
import re
from openpyxl import load_workbook


def _cell(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def _latest_sheet(wb):
    """시트 이름이 '202106(Rev.5)' 처럼 개정별이면 가장 최근 것."""
    names = wb.sheetnames
    live = [n for n in names if n.strip().startswith("실시간")]   # 담당자가 계속 고치는 현재 시트
    if live:
        return live[0]
    revs = []
    for n in names:
        m = re.search(r"Rev\.?\s*(\d+)", n)
        revs.append((int(m.group(1)) if m else -1, n))
    return max(revs)[1]


def _table(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[_latest_sheet(wb)]
    rows = [[_cell(v) for v in r] for r in ws.iter_rows(values_only=True)]
    for i, row in enumerate(rows[:8]):
        if row and row[0] == "No.":
            # '원료코드 Material Code' → '원료코드' (영문 설명은 뗀다)
            header = [re.sub(r"\s*[A-Za-z(].*$", "", c).strip() for c in row]
            return header, rows[i + 1:], ws.title
    raise ValueError("공급업체 목록의 머리행(No.)을 찾지 못했습니다: %s" % path)


def read_supplier_list(path):
    """[{공급업체명, 제조소 국가, 공급되는 품목, 원료코드?, 평가방법, 평가등급, 문서번호, 개정번호, 평가승인일}, ...]"""
    header, rows, sheet = _table(path)
    out = []
    for row in rows:
        if not row or not row[0]:
            continue
        rec = {h: (row[i] if i < len(row) else "") for i, h in enumerate(header) if h}
        rec["_sheet"] = sheet
        out.append(rec)
    return out


def find_supplier(records, code=None, item=None, name=None):
    """원료코드나 품목명(부분 일치)으로 한 줄 찾기."""
    for rec in records:
        if code and code in (rec.get("원료코드") or ""):
            return rec
    for rec in records:
        if item and item in (rec.get("공급되는 품목") or ""):
            if not name or name in (rec.get("공급업체명") or ""):
                return rec
    return None


def read_api_chain(path):
    """주성분 공급망 마스터파일 — {원료코드: {"api": 주성분명, "manufacturer": 제조소/국가, "chain": [1차, 2차, 3차]}}"""
    header, rows, _ = _table(path)
    out = {}
    for row in rows:
        if not row or not row[0]:
            continue
        rec = {h: (row[i] if i < len(row) else "") for i, h in enumerate(header) if h}
        code = rec.get("원료코드") or ""
        if code:
            chain = [rec.get(k, "") for k in header if "납품처" in k]
            out[code] = {"api": rec.get("주성분 명", ""), "manufacturer": rec.get("제조소/국가", ""),
                         "chain": [c for c in chain if c and c != "N/A"]}
    return out
