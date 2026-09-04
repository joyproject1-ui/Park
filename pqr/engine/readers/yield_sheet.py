# -*- coding: utf-8 -*-
"""7항 수율현황표를 읽는다.

회사 수율현황표는 제품마다 머리 부분이 다르다 — 어떤 것은 첫 줄이 바로 공정 이름이고, 어떤
것은 '연번 / 중요공정 별 수율 현황 (%)' 아래에 공정·기준·Lot No. 줄이 더 있고 Lot 은 둘째
칸이다(디겐타안연고). 자리를 정해 두면 다른 제품에서 값을 통째로 놓친다 — 실제로 디겐타
2026 자료에서 아홉 칸이 모두 '확인 필요' 로 나왔다. 그래서 자리를 정하지 않고 표를 보고 찾는다.
"""
import re

from openpyxl import load_workbook

# 제조번호: 영문과 숫자가 섞인 5~7 글자 (OGY301, OEY101 …). '조제'·'기준' 같은 말은 걸리지 않는다.
LOT = re.compile(r"^(?=.*[A-Z])(?=.*\d)[A-Z0-9]{5,7}$")


def _number(value):
    if value is None:
        return None
    try:
        return float(str(value).replace("%", "").replace(",", "").strip())
    except ValueError:
        return None


def _grid(path):
    ws = load_workbook(path, data_only=True, read_only=True).worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return [], 0
    width = max(len(r) for r in rows)
    return [r + [None] * (width - len(r)) for r in rows], width


def _lot_column(rows, width):
    """Lot 처럼 생긴 값이 가장 많은 열. 없으면 None."""
    hits = {}
    for row in rows:
        for c in range(width):
            if LOT.match(str(row[c] or "").strip()):
                hits[c] = hits.get(c, 0) + 1
    return max(sorted(hits), key=lambda c: hits[c]) if hits else None


def _stage_row(rows, value_cols, first_data):
    """공정 이름 줄 — 자료 줄 위쪽에서 값 칸에 숫자가 없는(기준 줄이 아닌) 가장 가까운 줄."""
    for i in range(first_data - 1, -1, -1):
        labels = [str(rows[i][c] or "").strip() for c in value_cols]
        if not any(labels):
            continue
        if any(re.search(r"\d", text) for text in labels):
            continue                       # '94% 이상', '86.5±6.5%' 같은 기준 줄
        return i
    return None


SPEC = re.compile(r"이상|이하|±|~")


def _layout(path):
    """(표, Lot 열, 자료 줄 번호들, 값 열들, {값 열: 공정 이름}). 못 읽으면 (None, ...)."""
    rows, width = _grid(path)
    lot_col = _lot_column(rows, width)
    if lot_col is None:
        return None, None, [], [], {}
    data_rows = [i for i, row in enumerate(rows) if LOT.match(str(row[lot_col] or "").strip())]
    value_cols = [c for c in range(lot_col + 1, width)
                  if any(_number(rows[i][c]) is not None for i in data_rows)]
    if not value_cols:
        return None, None, [], [], {}
    head = _stage_row(rows, value_cols, data_rows[0])
    names = {c: (str(rows[head][c] or "").strip() if head is not None else "") for c in value_cols}
    return rows, lot_col, data_rows, value_cols, names


def read_specs(path):
    """{공정명: 기준 글자} — 그해 수율현황표에 적힌 기준 줄. 없으면 {}.

    기준은 해가 바뀌며 개정된다(디겐타안연고 충전: 96.0 ± 3.5% → 86.5 ± 6.5%). 전년도 결재본에
    적힌 기준으로 견주면 멀쩡한 Lot 이 모두 '기준 벗어남' 으로 잡힌다 — 기준도 그해 자료에서 읽는다.
    """
    rows, lot_col, data_rows, value_cols, names = _layout(path)
    if rows is None:
        return {}
    for i in range(data_rows[0] - 1, -1, -1):
        texts = {c: str(rows[i][c] or "").strip() for c in value_cols}
        if any(SPEC.search(t) for t in texts.values()):
            return {names[c]: t for c, t in texts.items() if names.get(c) and t}
    return {}


def read_yields(path):
    """[(Lot, {공정명: 값문자열}), ...] — 값은 소수점 둘째 자리 문자열로 맞춘다."""
    rows, lot_col, data_rows, value_cols, names = _layout(path)
    if rows is None:
        return []
    out = []
    for i in data_rows:
        vals = {}
        for c in value_cols:
            name = names.get(c)
            if not name:
                continue
            number = _number(rows[i][c])
            if number is None:
                text = str(rows[i][c] or "").strip()
                if text:
                    vals[name] = text
            else:
                vals[name] = "%.2f" % number
        out.append((str(rows[i][lot_col]).strip(), vals))
    return out
