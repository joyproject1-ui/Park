# -*- coding: utf-8 -*-
"""ERP 에서 내려받은 표 — 주원료·자재 시험번호(.xls), 제조내역(PDF)."""
import os
import re

from ..pdftext import read_text, squash


def _sheet_rows(path):
    if path.lower().endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(path)
        if not book.codepage:
            # ERP 가 내려 준 .xls 에는 코드페이지 기록이 없어 xlrd 가 iso-8859-1 로 읽는다 —
            # 제품명이 'µð°ÕÅ¸¾È¿¬°í' 처럼 깨져 제품을 가려낼 수 없다.
            book = xlrd.open_workbook(path, encoding_override="cp949")
        sheet = book.sheet_by_index(0)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True, read_only=True).worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def test_no(raw):
    """ERP 표기 'R202405130070' → 보고서 표기 'R-2024-05-13-0070'."""
    s = str(raw or "").strip()
    m = re.match(r"^([A-Z])(\d{4})(\d{2})(\d{2})(\d{4})$", s)
    return "%s-%s-%s-%s-%s" % m.groups() if m else s


# ERP 가 그대로 내려 주는 표는 머리행이 없고 열이 열아홉이다. 쓰는 칸은 다음과 같다.
RAW_CODE, RAW_PRODUCT, RAW_TEST, RAW_LOT = 0, 7, 10, 16
CODE = re.compile(r"^[A-Z]{1,3}\d{3,6}$")                       # RBG201 · P17043
LOT = re.compile(r"^[A-Z0-9]{5,7}$")                            # OGY301


def _looks_raw(rows):
    """ERP 가 그대로 내려 준 표인지 — 머리행 없이 첫 칸이 원료 코드이고 열이 넉넉하다."""
    for row in rows[:3]:
        cells = [str(c or "").strip() for c in row]
        if len(cells) > RAW_LOT and CODE.match(cells[RAW_CODE]):
            return True
    return False


def read_material_tests(path, product=None):
    """원료/자재 코드 · 시험번호 · Lot 목록.  [(코드, 시험번호, Lot), ...] — Lot 없는 행은 뺀다.

    두 가지 꼴을 받는다.
      1) 사람이 추린 표: 머리행(원료 코드 · 시험번호 · Lot No.) + 세 칸
      2) ERP 가 그대로 내려 준 표: 머리행 없이 열아홉 칸. 담당자가 받는 것은 이쪽이다
         (2026-09 디겐타 자료). 이 꼴을 못 읽어 8.2 표에 지난해 값이 그대로 남았다.

    ERP 표는 원료 하나를 쓴 모든 제품이 함께 들어 있으므로(RSF101 은 후메론·톨론티 … 백여 줄)
    제품 이름으로 이 제품 줄만 가려낸다. Lot No. 는 열일곱째 칸에 그대로 있다 — 지어내지 않는다.
    """
    rows = _sheet_rows(path)
    if not _looks_raw(rows):
        out = []
        for row in rows[1:]:
            cells = [str(c or "").strip() for c in row[:3]]
            if len(cells) < 3 or not cells[2]:
                continue
            out.append((cells[0], test_no(cells[1]), cells[2]))
        return out

    key = re.sub(r"\s+", "", product or "")
    out = []
    for row in rows:
        cells = [str(c or "").strip() for c in row] + [""] * (RAW_LOT + 1)
        code, lot, test = cells[RAW_CODE], cells[RAW_LOT], cells[RAW_TEST]
        if not CODE.match(code) or not LOT.match(lot) or not test:
            continue                       # 제품에 쓰이지 않은 입고·시험 줄
        made = re.sub(r"\s+", "", cells[RAW_PRODUCT])
        if key and made != key:
            continue                       # 같은 원료를 쓰는 다른 제품 줄
        record = (code, test_no(test), lot)
        if record not in out:
            out.append(record)
    return out


def group_by_test(records):
    """같은 시험번호를 쓴 Lot 을 묶는다 (보고서 8.2 표의 세로 병합 단위). 순서는 첫 등장 순."""
    order, groups = [], {}
    for code, test, lot in records:
        key = (code, test)
        if key not in groups:
            groups[key] = []
            order.append(key)
        if lot not in groups[key]:
            groups[key].append(lot)
    return [(code, test, groups[(code, test)]) for code, test in order]


LOT_LINE = re.compile(r"^\s*([A-Z]{2}[A-Z0-9]{4})\s+(\S.*?)\s+(\d{4}\.\d{2}\.\d{2})\s+(\d{4}\.\d{2}\.\d{2})\s*$", re.M)


def read_manufacturing(path):
    """제조내역 PDF — [(Lot, 품명, 제조일자, 사용기한), ...]"""
    text = squash(read_text(path))
    return [(m.group(1), m.group(2).strip(), m.group(3), m.group(4)) for m in LOT_LINE.finditer(text)]
