# -*- coding: utf-8 -*-
"""이미 채워 둔 안정성 경향 분석 결과(HLF-QC-126-06)를 다시 읽는다.

담당자는 해마다 이 파일을 이어서 쓴다 — 지난해 파일에 새 시점을 덧붙이는 식이다
(담당자 2026-09: "16. 안정성시험 경향표를 13항 안정성 시험 자료를 참고해서 최신 파일로
신규 작성"). 지난 파일의 값은 담당자가 손으로 옮겨 적어 둔 것이라 스캔 판독보다 믿을 만하므로,
그것을 바탕으로 삼고 올해 시험일지에서 읽은 새 시점만 덧붙인다.
"""
from __future__ import unicode_literals

import os
import re

from openpyxl import load_workbook

POINTS = ["Initial", "3M", "6M", "9M", "12M", "18M", "24M", "36M", "48M", "60M"]
FIRST_ROW, ROWS = 38, 30
FIRST_COL = 3                    # C 열 = Initial
TITLE = "안정성 시험 경향 분석"
LOT = re.compile(r"^[A-Z0-9]{5,8}$")


def _text(value):
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip()


def _number(value):
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"-?\d+(?:\.\d+)?", str(value or ""))
    return float(m.group()) if m else None


def is_trend_file(path):
    """채워 둔 경향 분석 파일인가 — 제목이 맞고 제품명(C3)이 적혀 있어야 한다."""
    if not str(path or "").lower().endswith(".xlsx") or os.path.basename(str(path)).startswith("~$"):
        return False
    try:
        book = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return False
    try:
        for sheet in book.worksheets:
            head = " ".join(_text(sheet.cell(1, c).value) for c in range(1, 12))
            if TITLE in head and _text(sheet["C3"].value):
                return True
        return False
    finally:
        book.close()


def read_trend(path):
    """[{"sheet", "item", "product", "lcl", "ucl", "storage", "lots": [(제조번호, {시점: 값})]}]

    값이 하나도 없는 시트(쓰지 않은 pH·삼투압 …)는 건너뛴다.
    """
    book = load_workbook(path, data_only=True, read_only=True)
    out = []
    try:
        for sheet in book.worksheets:
            head = " ".join(_text(sheet.cell(1, c).value) for c in range(1, 12))
            if TITLE not in head:
                continue
            lots = []
            for i in range(ROWS):
                row = FIRST_ROW + i
                lot = _text(sheet.cell(row, 2).value)
                if not lot or not LOT.match(lot.upper()):
                    continue
                values = {}
                for k, point in enumerate(POINTS):
                    got = _number(sheet.cell(row, FIRST_COL + k).value)
                    if got is not None:
                        values[point] = got
                if values:
                    lots.append((lot, values))
            if not lots:
                continue
            out.append({"sheet": sheet.title, "item": _text(sheet["G3"].value),
                        "product": _text(sheet["C3"].value),
                        "storage": _text(sheet["C4"].value),
                        "lcl": _number(sheet["H4"].value), "ucl": _number(sheet["J4"].value),
                        "lots": lots})
    finally:
        book.close()
    return out


def component_of(item):
    """시험항목 글에서 성분 이름 — '함량 - 플루오로메톨론(%)' → '플루오로메톨론'."""
    text = re.sub(r"\(.*?\)", "", _text(item))
    parts = re.split(r"[-–—:]", text, 1)
    return (parts[1] if len(parts) > 1 else parts[0]).strip()
