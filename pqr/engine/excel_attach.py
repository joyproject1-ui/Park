# -*- coding: utf-8 -*-
"""보고서에 딸린 엑셀 — Cpk 계산 파일(HLF-QC-126-08/09) 과 안정성 경향 분석(HLF-QC-126-06).

Cpk 파일은 전년도 결재본 압축에 든 것을 그대로 물려받아 값만 갈아 끼운다(수식·서식 보존).
안정성 경향 파일은 서식(HLF-QC-126-06)이 제품 폴더·입력 폴더의 '서식' 폴더·프로그램에
있으면 채운다.
"""
import os
import re
import shutil
import tempfile
import zipfile

from openpyxl import load_workbook

from . import convert
from . import stability_xlsx
from .readers import trend as trend_reader
from . import xls_fill

CPK_ITEMS = [   # (파일 이름에 든 낱말, 완제 성적서 항목)
    ("금속성이물(개개)", "metal_each"), ("금속성이물(합계)", "metal_total"),
    ("입자도", "particle"), ("함량", "assay"),
]


def _num(v):
    m = re.search(r"-?\d+(?:\.\d+)?", str(v or ""))
    return float(m.group()) if m else None


def _previous_xls(previous_path, workdir):
    """결재본(.zip 또는 폴더 옆)의 Cpk .xls 들. {낱말: 경로}"""
    found = {}
    cands = []
    if previous_path and previous_path.lower().endswith(".zip"):
        with zipfile.ZipFile(previous_path) as z:
            for info in z.infolist():
                name = info.filename
                try:
                    name = name.encode("cp437").decode("cp949")
                except Exception:
                    pass
                if name.lower().endswith((".xls", ".xlsx")) and not os.path.basename(name).startswith("~$"):
                    target = os.path.join(workdir, os.path.basename(name))
                    with open(target, "wb") as h:
                        h.write(z.read(info.filename))
                    cands.append(target)
    else:
        folder = os.path.dirname(previous_path) if previous_path else ""
        if folder and os.path.isdir(folder):
            cands = [os.path.join(folder, n) for n in os.listdir(folder) if n.lower().endswith((".xls", ".xlsx"))]
    for path in cands:
        for word, _ in CPK_ITEMS:
            if word in os.path.basename(path) and "Cpk" in os.path.basename(path):
                found[word] = path
    return found


def _fill_cpk_xlsx(src_xls, dst_xlsx, values, today):
    """Excel·LibreOffice 가 없을 때의 대비책 — .xlsx 로 바꿔 값만 채운다(그래프는 사라진다)."""
    tmp = dst_xlsx + ".tmp.xlsx"
    convert.to_xlsx(src_xls, tmp)
    wb = load_workbook(tmp)
    ws = wb.worksheets[0]
    ws["K4"] = today
    for i in range(35):
        ws.cell(row=10 + i, column=2).value = values[i] if i < len(values) else None
    wb.save(dst_xlsx)
    os.remove(tmp)
    return dst_xlsx


def fill_cpk(src_xls, dst_xls, values, today):
    """B10~B44 에 결과값, K4 에 작성일. 서식의 수식·그래프를 그대로 둔 채 .xls 로 저장한다."""
    xls_fill.fill(src_xls, dst_xls, {"K4": today}, values)
    return dst_xls


def write_cpk_files(folder, data, previous_path, today, lots=None):
    """내수용 Lot 의 완제 성적서 값으로 Cpk 파일 4종을 제품 폴더에 만든다. [(이름, 경로)]"""
    lots = lots or data.domestic
    from . import qc
    if not qc.cpk_applies(len(lots)):
        # QC-126: 평가 년도 생산 Lot 이 기준(10) 미만이면 Cpk 를 산출하지 않는다 — 계산 파일도 만들지 않는다
        data.issues.append(("첨부", "", "평가 년도 생산 %d Lot 으로 %d Lot 미만 — QC-126 에 따라 Cpk 계산 파일을 만들지 않음"
                            % (len(lots), qc.cpk_min_lots())))
        return []
    work = tempfile.mkdtemp(prefix="pqr-cpk-")
    sources = _previous_xls(previous_path, work)
    out = []
    for word, key in CPK_ITEMS:
        src = sources.get(word)
        if not src:
            data.issues.append(("첨부", "", "전년도 결재본에 '%s Cpk 계산 파일' 이 없어 만들지 못함" % word))
            continue
        vals = [_num((data.coa.get(l) or {}).get("924", {}).get(key)) for l in lots]
        vals = [v for v in vals if v is not None]
        name = os.path.basename(src)
        dst = os.path.join(folder, name)
        try:
            fill_cpk(src, dst, vals, today)
            out.append((name, dst))
        except xls_fill.FillError as error:
            name = re.sub(r"\.xls$", ".xlsx", name)
            dst = os.path.join(folder, name)
            try:                                  # 그래프는 잃지만 값이라도 남긴다
                _fill_cpk_xlsx(src, dst, vals, today)
                out.append((name, dst))
                data.issues.append(("첨부", name, "%s — 그래프 없이 값만 채웠습니다" % error))
            except Exception as second:
                data.issues.append(("첨부", name, "Cpk 파일을 만들지 못함: %s" % second))
        except Exception as error:
            data.issues.append(("첨부", name, "Cpk 파일을 만들지 못함: %s" % error))
    shutil.rmtree(work, ignore_errors=True)
    return out


def _is_blank_form(path):
    """제품명 칸(C3)이 비어 있으면 아직 채우지 않은 서식이다.

    담당자가 올리는 서식 파일 이름에도 '결과'가 들어 있어(‘HLF-QC-126-06 안정성 시험 경향
    분석 결과 (Rev.001).xlsx’) 이름만으로는 채운 파일과 가릴 수 없다 — 안을 보고 가린다.
    """
    try:
        book = load_workbook(path, data_only=True, read_only=True)
        try:
            return not str(book.worksheets[0]["C3"].value or "").strip()
        finally:
            book.close()
    except Exception:
        return False


def _find_stability_form(folder, input_dir, product_dir=None):
    for base in (product_dir, folder, os.path.join(input_dir or "", "서식"), input_dir or "",
                 os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")):
        if not base or not os.path.isdir(base):
            continue
        found = [os.path.join(base, n) for n in sorted(os.listdir(base))
                 if n.lower().endswith(".xlsx") and not n.startswith("~$")
                 and ("12606" in n.replace("-", "") or "126-06" in n)]
        blank = [p for p in found if _is_blank_form(p)]
        if blank:
            return blank[0]
    return None


COL = {"Initial": "C", "초기": "C", "3M": "D", "6M": "E", "9M": "F", "12M": "G", "18M": "H", "24M": "I",
       "36M": "J", "48M": "K", "60M": "L"}


W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def written_by(report_path):
    """보고서 '검토 및 승인' 표에서 작 성 (Written by) 의 성 명을 읽는다.

    첨부 엑셀의 작성자를 본문과 같게 맞추기 위한 것이다. 못 찾으면 빈 문자열.
    """
    if not report_path or not os.path.isfile(report_path):
        return ""
    try:
        from lxml import etree
        with zipfile.ZipFile(report_path) as z:
            root = etree.fromstring(z.read("word/document.xml"))
    except Exception:
        return ""

    def text(el):
        return "".join(t.text or "" for t in el.iter(W + "t"))

    for tbl in root.iter(W + "tbl"):
        rows = tbl.findall(W + "tr")
        for i, tr in enumerate(rows):
            if "Written by" not in text(tr):
                continue
            for nxt in rows[i + 1:]:
                cells = nxt.findall(W + "tc")
                if len(cells) < 2:
                    continue
                name = re.sub(r"\s+", "", text(cells[1]))
                if name:
                    return name
            return ""
    return ""


def _grouped(points):
    """{Lot: {시점: 값}} 또는 {포장구분: {Lot: {시점: 값}}} 을 [(구분, [(Lot, 값)])] 로 편다."""
    if not points:
        return []
    inner = list(points.values())
    if inner and isinstance(inner[0], dict) and inner[0] and \
            all(isinstance(v, dict) for v in inner[0].values()):
        return [(k, list(v.items())) for k, v in points.items()]
    return [("", list(points.items()))]


def write_stability_workbook(folder, data, product, today, input_dir=None, report_path=None,
                             product_dir=None):
    """HLF-QC-126-06 — 시점별 함량을 채운다. 포장 규격이 나뉘어 있으면 파일도 나눈다.

    작성자(M3)·작성일(M4)은 보고서 본문의 작성자·작성일자와 같게 넣는다.
    반환값은 [(파일명, 경로), ...] 이며 만들지 못하면 빈 목록이다.
    """
    stab = getattr(data, "stability", None)
    points = (stab or {}).get("points") or {}
    logs = getattr(data, "stability_logs", None)
    form = _find_stability_form(folder, input_dir, product_dir)
    if not form:
        data.issues.append(("첨부", "", "안정성 경향 분석 서식(HLF-QC-126-06)을 찾지 못해 만들지 못함 — 제품 폴더나 입력 폴더의 '서식' 폴더에 두세요"))
        return []
    if logs or getattr(data, "stability_trend", None):
        return _write_trend(form, folder, data, product, today, report_path)
    if not points:
        data.issues.append(("첨부", "", "안정성 시험일지 판독값이 없어 경향 분석 파일을 만들지 못함 — 시험일지를 올리거나 담당자가 직접 기입"))
        return []
    base = product.get("name") or ""
    author = written_by(report_path)
    made = []
    for label, lots in _grouped(points):
        title = base if not label or label in base else "%s(%s)" % (base, label)
        name = "HLF-QC-126-06 안정성 시험 경향 분석 결과 - %s.xlsx" % title
        dst = os.path.join(folder, name)
        lots = [(lot, {k: _num(v) for k, v in vals.items()}) for lot, vals in lots]
        stability_xlsx.build(form, dst, title, lots,
                             lcl=(stab or {}).get("lcl", 90),
                             ucl=(stab or {}).get("ucl", 110),
                             prepared_by=author, prepared_on=today)
        made.append((name, dst))
    return made


FORM_SHEETS = ("함량", "A", "B", "기타", "총", "pH", "삼투압")


def _assay_limits(data):
    """{성분: (하한, 상한)} — 완제 성적서의 함량 규격. 제품마다 다르므로 여기서 읽는다."""
    out = {}
    for lot in (data.coa or {}).values():
        for a in (lot.get("924") or {}).get("assays") or []:
            part = (a.get("part") or "").strip()
            if part and part not in out:
                lo, hi = _num(a.get("lo")), _num(a.get("hi"))
                if lo is not None and hi is not None:
                    out[part] = (lo, hi)
    return out


def _same_part(a, b):
    a, b = re.sub(r"\s+", "", a or ""), re.sub(r"\s+", "", b or "")
    return bool(a) and bool(b) and (a in b or b in a)


def _write_trend(form, folder, data, product, today, report_path):
    """안정성 경향 분석(HLF-QC-126-06)을 새로 만든다 — 지난 경향표 + 올해 시험일지.

    담당자 지시(2026-09): "16. 안정성시험 경향표를 13항 안정성 시험 자료를 참고해서 최신
    파일로 신규 작성." 지난 경향표의 값은 담당자가 옮겨 적어 둔 것이라 그대로 이어받고,
    올해 시험일지에서 읽은 시점만 덧붙인다. 평가 기간을 넘어선 시점(다음 해에 끝난 시험)은
    넣지 않는다 — 보고서 13.3 의 경향 범위와 같은 값이어야 한 벌의 자료로 읽힌다.
    """
    seed = list(getattr(data, "stability_trend", None) or [])
    logs = list(getattr(data, "stability_logs", None) or [])
    year_to = int((getattr(data, "period", None) or {}).get("to") or 0) or None
    limits = _assay_limits(data)

    def points_of(one, part):
        out = {}
        for point in one.get("points", []):
            got = re.findall(r"\d{4}", point.get("done") or "")
            if year_to and got and int(got[0]) > year_to:
                continue
            value = _num((point.get("assays") or {}).get(part))
            if value is not None:
                out[point["period"]] = float(value)
        return out

    parts = []                                    # [(성분, 시험항목 글, 하한, 상한, 지난 Lot)]
    for sheet in seed:
        part = trend_reader.component_of(sheet.get("item"))
        lo, hi = limits.get(next((k for k in limits if _same_part(k, part)), ""), (None, None))
        parts.append((part, sheet.get("item") or ("함량 - %s(%%)" % part),
                      lo if lo is not None else sheet.get("lcl"),
                      hi if hi is not None else sheet.get("ucl"),
                      list(sheet.get("lots") or [])))
    if not parts:                                 # 지난 경향표가 없는 첫해 — 성적서·시험일지에서
        names = list(limits)
        if not names:
            names = [part for one in logs for point in one.get("points", [])
                     for part in (point.get("assays") or {})]
            names = list(dict.fromkeys(names))
        for part in names:
            lo, hi = limits.get(part, (90, 110))
            parts.append((part, "함량 - %s(%%)" % part, lo, hi, []))
    if not parts:
        data.issues.append(("첨부", "", "안정성 경향 분석에 넣을 성분을 찾지 못해 만들지 못함 — "
                                        "지난 경향표(HLF-QC-126-06)나 시험일지 판독값이 필요합니다"))
        return []

    sheets, added = [], 0
    for i, (part, item, lo, hi, old_lots) in enumerate(parts[:len(FORM_SHEETS)]):
        rows = []
        seen = {}
        for lot, values in old_lots:              # 지난 경향표의 차례를 지킨다
            seen[lot] = dict(values)
            rows.append(lot)
        for one in logs:
            values = points_of(one, part)
            if not values:
                continue
            lot = one["lot"]
            if lot not in seen:
                seen[lot] = {}
                rows.append(lot)
            for period, value in values.items():
                if seen[lot].get(period) != value:
                    added += 1
                seen[lot][period] = value
        sheets.append({"form_sheet": FORM_SHEETS[i], "name": "함량(%s)" % part, "item": item,
                       "lots": [(lot, seen[lot]) for lot in rows],
                       "lcl": float(lo if lo is not None else 90),
                       "ucl": float(hi if hi is not None else 110)})
    if not any(sheet["lots"] for sheet in sheets):
        data.issues.append(("첨부", "", "안정성 시험 결과값이 없어 경향 분석 파일을 만들지 못함 — "
                                        "13항에 시험일지 판독값(.json)이나 지난 경향표를 두세요"))
        return []
    if not added:
        data.issues.append(("첨부", "", "13항 시험일지에서 새 시점을 읽지 못해 지난 경향표 값만 "
                                        "옮겼습니다 — 올해 시점을 직접 채우세요"))

    name = "HLF-QC-126-06 안정성 시험 경향 분석 결과 - %s.xlsx" % (product.get("name") or "")
    dst = os.path.join(folder, name)
    stability_xlsx.build_multi(form, dst, product.get("name") or "", sheets,
                               prepared_by=written_by(report_path), prepared_on=today)
    return [(name, dst)]
