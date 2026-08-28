"""마스터파일에서 해당 제품 몫만 골라냅니다.

밸리데이션·적격성 자료는 전 제품이 한 파일에 들어 있고, 제품 열이 없는 경우가
많습니다. 대신 **공정밸리데이션(PV) 마스터파일에는 제품명이 있고, 그 제품이
어느 시트(=제조 라인)에 있는지가 곧 라인**이므로, 그 라인으로 설비 목록을 거릅니다.

    제품명 → (PV 마스터의 시트) → 라인 → 그 라인의 설비

병합된 칸은 빈칸으로 읽히므로, 위 행의 값을 이어받아 채웁니다.
"""

import re

from .schema import parse_date
from .tabular import read_xlsx


class MasterError(Exception):
    """마스터파일에서 필요한 표를 찾지 못했을 때."""


def _squash(text):
    return re.sub(r"\s+", "", str(text or "")).lower()


def _read_sheet(path, sheet):
    """openpyxl 없이 표준 라이브러리 리더로 시트를 행 목록(원본 순서)으로 읽습니다."""
    return read_xlsx(path, sheet=sheet)


def find_header(rows, required, limit=12):
    """머리글 행의 위치와 {열이름: 인덱스} 를 찾습니다.

    rows 는 열 이름을 모르는 상태이므로, 위쪽 몇 줄을 훑어 필요한 낱말이
    모두 들어 있는 줄을 머리글로 봅니다.
    """
    for index, row in enumerate(rows[:limit]):
        cells = [_squash(value) for value in row]
        if all(any(_squash(word) in cell for cell in cells) for word in required):
            return index, cells
    raise MasterError("머리글을 찾지 못했습니다 (찾는 낱말: %s)" % ", ".join(required))


def _column(cells, *words):
    for index, cell in enumerate(cells):
        if any(_squash(word) in cell for word in words):
            return index
    return None


def _carry(rows, columns):
    """병합으로 비어 있는 칸을 위 행의 값으로 채웁니다."""
    carried = []
    last = {}
    for row in rows:
        filled = list(row)
        for index in columns:
            if index is None or index >= len(filled):
                continue
            value = str(filled[index]).strip() if filled[index] is not None else ""
            if value:
                last[index] = value
            else:
                filled[index] = last.get(index, "")
        carried.append(filled)
    return carried


# --------------------------------------------------------------------------
# 10.1 공정밸리데이션 (PV) 마스터파일
# --------------------------------------------------------------------------

def parse_pv_sheet(rows, product_name):
    """PV 마스터 시트 한 장에서 제품의 밸리데이션 이력을 뽑습니다.

    제품이 없으면 None. 계획서 한 건에 검증 Lot 여러 개가 딸리고,
    Lot 마다 보고서가 따로 있는 형태입니다.
    """
    try:
        header_index, cells = find_header(rows, ("제품명", "lot"))
    except MasterError:
        return None
    col = {
        "no": _column(cells, "no."),
        "product": _column(cells, "제품명"),
        "plan": _column(cells, "계획서"),
        "reason": _column(cells, "사유"),
        "kind": _column(cells, "종류"),
        "seq": _column(cells, "lot.no", "lotno"),
        "report": _column(cells, "보고서"),
        "year": _column(cells, "재밸리데이션"),
        "note": _column(cells, "비고"),
    }
    if col["product"] is None:
        return None
    body = _carry(rows[header_index + 1:],
                  [col["no"], col["product"], col["plan"], col["reason"],
                   col["kind"], col["year"]])
    matched = [row for row in body
               if col["product"] < len(row)
               and _squash(product_name) in _squash(row[col["product"]])]
    if not matched:
        return None

    entries = []
    index = {}
    for row in matched:
        def cell(key):
            position = col[key]
            if position is None or position >= len(row) or row[position] is None:
                return ""
            return str(row[position]).strip()

        # 계획서 한 건이 하나의 밸리데이션입니다. 보고서는 Lot 마다 달릴 수 있습니다.
        plan = cell("plan")
        if plan not in index:
            index[plan] = {
                "plan": plan, "reason": cell("reason"), "kind": cell("kind"),
                "revalidation_year": cell("year"), "note": cell("note"),
                "reports": [], "lots": [],
            }
            entries.append(index[plan])
        entry = index[plan]
        report = cell("report")
        if report and report not in entry["reports"]:
            entry["reports"].append(report)
        lot = _lot_of(row, col["seq"])
        if lot:
            lot["report"] = report
            entry["lots"].append(lot)
    return entries


def read_pv_master(path, product_name, sheets=None):
    """제품이 실린 시트를 찾아 PV 이력을 돌려줍니다.

    반환값: {"line": 시트이름, "entries": [...]}  (시트 이름이 곧 제조 라인)
    """
    from openpyxl import load_workbook                     # 마스터파일 전용(선택 의존성)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in (sheets or workbook.sheetnames):
            rows = [list(row) for row in
                    workbook[sheet].iter_rows(max_col=20, values_only=True)]
            entries = parse_pv_sheet(rows, product_name)
            if entries:
                return {"line": sheet.strip(), "entries": entries}
    finally:
        workbook.close()
    raise MasterError("PV 마스터파일에서 '%s' 를 찾지 못했습니다." % product_name)


def _lot_of(row, seq_index):
    """'1st' 다음 칸의 제조번호와 제조일자를 집어냅니다."""
    if seq_index is None or seq_index >= len(row):
        return None
    seq = str(row[seq_index] or "").strip()
    if not seq:
        return None
    lot_no, mfg_date = "", ""
    for value in row[seq_index + 1:]:
        text = str(value or "").strip()
        if not text:
            continue
        if not lot_no:
            lot_no = text
            continue
        parsed = parse_date(text)
        if parsed:
            mfg_date = parsed.isoformat()
        break
    return {"seq": seq, "lot_no": lot_no, "mfg_date": mfg_date}


# --------------------------------------------------------------------------
# 10.2 · 10.5 설비 적격성 마스터파일
# --------------------------------------------------------------------------

def parse_equipment_sheet(rows, line):
    """설비 마스터 시트 한 장에서 라인에 속한 설비의 최신 이력을 뽑습니다."""
    try:
        header_index, cells = find_header(rows, ("관리번호", "설비명"))
    except MasterError:
        return []
    col = {
        "dept": _column(cells, "부서"),
        "line": _column(cells, "라인"),
        "asset_id": _column(cells, "관리번호"),
        "asset": _column(cells, "설비명"),
        "doc": _column(cells, "문서번호"),
        "rev": _column(cells, "revno", "rev no"),
        "approved": _column(cells, "승인일자", "승인일"),
        "year": _column(cells, "재적격성"),
        "reason": _column(cells, "개정사유", "사유"),
    }
    body = _carry(rows[header_index + 1:],
                  [col["dept"], col["line"], col["asset_id"], col["asset"]])

    latest = {}
    for row in body:
        def cell(key):
            index = col[key]
            if index is None or index >= len(row) or row[index] is None:
                return ""
            return str(row[index]).strip()

        if line and _squash(line) not in _squash(cell("line")):
            continue
        asset_id = cell("asset_id")
        if not asset_id or not cell("asset"):
            continue
        approved = parse_date(cell("approved"))
        record = {
            "asset_id": asset_id, "asset": cell("asset"),
            "dept": cell("dept"), "line": cell("line"),
            "doc_no": cell("doc"), "rev": cell("rev"),
            "approved": approved.isoformat() if approved else cell("approved"),
            "revalidation_year": cell("year"), "reason": cell("reason"),
        }
        previous = latest.get(asset_id)
        if previous is None or (approved and (
                not previous.get("approved") or record["approved"] > previous["approved"])):
            latest[asset_id] = record
    return sorted(latest.values(), key=lambda item: item["asset_id"])


def read_equipment_master(path, line, sheet=None):
    """라인에 속한 설비의 최신 적격성 이력을 돌려줍니다.

    설비 하나에 개정 이력이 여러 행 있으므로, 승인일자가 가장 늦은 행을 씁니다.
    """
    from openpyxl import load_workbook
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for name in ([sheet] if sheet else workbook.sheetnames):
            rows = [list(row) for row in
                    workbook[name].iter_rows(max_col=20, values_only=True)]
            found = parse_equipment_sheet(rows, line)
            if found:
                return found
    finally:
        workbook.close()
    raise MasterError("설비 마스터파일에서 라인 '%s' 를 찾지 못했습니다." % line)
